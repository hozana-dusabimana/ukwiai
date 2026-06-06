from __future__ import annotations
import os
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
from typing import Any

from sqlalchemy import select, func
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)
from openpyxl import Workbook

from app.core.config import settings
from app.models.project import Project
from app.models.budget import BudgetRecord
from app.models.analysis import ProgressAnalysis
from app.models.cost import CostEstimation
from app.models.report import Report
from app.services.cost_estimation import (
    total_recorded_expenses,
    total_ai_inferred_cost,
    compute_cost_estimation,
)


def _ensure_reports_dir() -> Path:
    p = Path(settings.REPORTS_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _filename(project: Project, report_type: str, ext: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    code = (project.project_code or f"P{project.id}").replace("/", "_")
    return f"{code}_{report_type}_{ts}.{ext}"


def _gather(db: Session, project: Project) -> dict[str, Any]:
    expenses = db.scalars(
        select(BudgetRecord).where(BudgetRecord.project_id == project.id).order_by(BudgetRecord.expense_date)
    ).all()
    total_spent = total_recorded_expenses(db, project.id)
    ai_inferred = total_ai_inferred_cost(db, project.id)
    effective_spent = total_spent if total_spent > ai_inferred else ai_inferred
    latest_analysis = db.scalars(
        select(ProgressAnalysis)
        .where(ProgressAnalysis.project_id == project.id)
        .order_by(ProgressAnalysis.analysis_date.desc())
        .limit(1)
    ).first()
    # Compute a fresh estimation (not persisted) so the report always reflects
    # the current effective spend + latest progress, rather than rendering a
    # stale stored snapshot whose projected total may predate logged costs.
    progress = (
        latest_analysis.predicted_progress_percentage
        if latest_analysis and latest_analysis.predicted_progress_percentage is not None
        else 0
    )
    latest_estimation = compute_cost_estimation(db, project, progress, persist=False)
    by_category = dict(
        db.execute(
            select(BudgetRecord.expense_category, func.coalesce(func.sum(BudgetRecord.amount), 0))
            .where(BudgetRecord.project_id == project.id)
            .group_by(BudgetRecord.expense_category)
        ).all()
    )
    return {
        "expenses": expenses,
        "total_spent": total_spent,
        "ai_inferred_spent": ai_inferred,
        "effective_spent": effective_spent,
        "latest_analysis": latest_analysis,
        "latest_estimation": latest_estimation,
        "by_category": by_category,
    }


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("UkwiTitle", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#1f4e79")))
    styles.add(ParagraphStyle("UkwiH2", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#1f4e79")))
    return styles


def generate_pdf_report(db: Session, project: Project, generated_by_id: int, report_type: str = "full") -> Report:
    data = _gather(db, project)
    out_dir = _ensure_reports_dir()
    fname = _filename(project, report_type, "pdf")
    fpath = out_dir / fname

    doc = SimpleDocTemplate(str(fpath), pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm)
    styles = _pdf_styles()
    elems = []

    elems.append(Paragraph("UKWI Company Ltd — Construction Monitoring Report", styles["UkwiTitle"]))
    elems.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elems.append(Spacer(1, 0.4 * cm))

    info = [
        ["Project", project.project_name],
        ["Code", project.project_code],
        ["Location", project.location or "-"],
        ["Client", project.client_name or "-"],
        ["Status", project.status.value],
        ["Budget", f"{project.total_budget:,.2f}"],
        ["Start", str(project.start_date or "-")],
        ["Expected end", str(project.expected_end_date or "-")],
    ]
    t = Table(info, colWidths=[5 * cm, 11 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#dee9f5")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    elems += [Paragraph("Project Information", styles["UkwiH2"]), t, Spacer(1, 0.5 * cm)]

    if report_type in ("progress", "full"):
        elems.append(Paragraph("AI Progress Analysis", styles["UkwiH2"]))
        a = data["latest_analysis"]
        if a:
            rows = [
                ["Predicted stage", a.predicted_stage or "-"],
                ["Predicted progress", f"{a.predicted_progress_percentage or 0}%"],
                ["Confidence", f"{(a.confidence_score or 0) * 100:.1f}%"],
                ["Model version", a.model_version or "-"],
                ["Date", a.analysis_date.strftime("%Y-%m-%d %H:%M")],
            ]
            t = Table(rows, colWidths=[5 * cm, 11 * cm])
            t.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                                   ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey)]))
            elems.append(t)
        else:
            elems.append(Paragraph("No AI analysis recorded yet.", styles["Normal"]))
        elems.append(Spacer(1, 0.5 * cm))

    if report_type in ("budget", "full"):
        elems.append(Paragraph("Budget Summary", styles["UkwiH2"]))
        e = data["latest_estimation"]
        rows = [
            ["Total budget", f"{project.total_budget:,.2f}"],
            ["Recorded spend", f"{data['total_spent']:,.2f}"],
            ["AI-inferred spend", f"{data['ai_inferred_spent']:,.2f}"],
            ["Effective spend", f"{data['effective_spent']:,.2f}"],
            ["Estimated spend (AI)", f"{e.estimated_cost_used if e else 0:,.2f}"],
            ["Variance", f"{e.variance if e else 0:,.2f}"],
            ["Projected total", f"{e.projected_total_cost if e else 0:,.2f}"],
            ["Deviation", e.deviation_status.value if e else "-"],
        ]
        t = Table(rows, colWidths=[5 * cm, 11 * cm])
        t.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                               ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey)]))
        elems += [t, Spacer(1, 0.4 * cm)]

        if data["by_category"]:
            elems.append(Paragraph("Spend by category", styles["UkwiH2"]))
            cat_rows = [["Category", "Amount"]]
            for cat, amt in data["by_category"].items():
                cat_rows.append([str(cat).title() if hasattr(cat, "value") is False else cat.value.title(), f"{Decimal(str(amt)):,.2f}"])
            tt = Table(cat_rows, colWidths=[5 * cm, 11 * cm])
            tt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ]))
            elems += [tt, Spacer(1, 0.4 * cm)]

    if report_type == "full":
        elems.append(PageBreak())
        elems.append(Paragraph("Expense Detail", styles["UkwiH2"]))
        rows = [["Date", "Category", "Amount", "Description"]]
        for x in data["expenses"]:
            rows.append([
                str(x.expense_date),
                x.expense_category.value,
                f"{x.amount:,.2f}",
                (x.description or "")[:60],
            ])
        if len(rows) == 1:
            rows.append(["-", "-", "-", "No expenses recorded"])
        t = Table(rows, colWidths=[3 * cm, 3 * cm, 3 * cm, 7 * cm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elems.append(t)

    doc.build(elems)

    rep = Report(
        project_id=project.id,
        report_type=report_type,
        file_path=str(fpath),
        generated_by=generated_by_id,
        period_start=None,
        period_end=None,
    )
    db.add(rep)
    db.flush()
    return rep


def generate_excel_report(db: Session, project: Project, generated_by_id: int, report_type: str = "budget") -> Report:
    data = _gather(db, project)
    out_dir = _ensure_reports_dir()
    fname = _filename(project, report_type, "xlsx")
    fpath = out_dir / fname

    wb = Workbook()
    ws = wb.active
    ws.title = "Project"
    ws.append(["UKWI Construction Monitor"])
    ws.append([])
    ws.append(["Project", project.project_name])
    ws.append(["Code", project.project_code])
    ws.append(["Status", project.status.value])
    ws.append(["Budget", float(project.total_budget or 0)])
    ws.append(["Recorded spend", float(data["total_spent"])])
    ws.append(["AI-inferred spend", float(data["ai_inferred_spent"])])
    ws.append(["Effective spend", float(data["effective_spent"])])

    e = data["latest_estimation"]
    if e:
        ws.append(["Estimated spend (AI)", float(e.estimated_cost_used or 0)])
        ws.append(["Variance", float(e.variance or 0)])
        ws.append(["Projected total", float(e.projected_total_cost or 0)])
        ws.append(["Deviation", e.deviation_status.value])

    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Date", "Category", "Amount", "Description", "Stage ID"])
    for x in data["expenses"]:
        ws2.append([
            x.expense_date,
            x.expense_category.value,
            float(x.amount),
            x.description or "",
            x.stage_id or "",
        ])

    ws3 = wb.create_sheet("By Category")
    ws3.append(["Category", "Total"])
    for cat, amt in data["by_category"].items():
        cat_name = cat.value if hasattr(cat, "value") else str(cat)
        ws3.append([cat_name, float(amt or 0)])

    wb.save(str(fpath))

    rep = Report(
        project_id=project.id,
        report_type=report_type,
        file_path=str(fpath),
        generated_by=generated_by_id,
    )
    db.add(rep)
    db.flush()
    return rep
